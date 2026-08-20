"""
reporting_pipeline.py
End-to-end automated weekly ops report:
  raw CSV  ->  clean & validate  ->  compute KPIs  ->  formatted Excel report
           ->  LLM writes a plain-English executive summary

Run:  python reporting_pipeline.py
Optional (for the real AI summary): set GEMINI_API_KEY first.
"""
import os
import re
import json
import urllib.request
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

RAW_FILE = "weekly_ops_raw.csv"
REPORT_FILE = "weekly_ops_report.xlsx"
SUMMARY_FILE = "executive_summary.txt"


# ---------------------------------------------------------------- 1. CLEAN
def clean(df: pd.DataFrame) -> pd.DataFrame:
    before = len(df)
    df = df.drop_duplicates()                                   # remove dupes
    df["industry"] = df["industry"].str.strip()                # trim whitespace
    df["plan_type"] = df["plan_type"].str.strip().str.capitalize()
    df["payment_status"] = (
        df["payment_status"].astype(str).str.strip().str.capitalize().replace("", pd.NA)
    )
    # mrr: strip $ and commas, coerce to number, fill blanks with 0
    df["mrr"] = (
        df["mrr"].astype(str).str.replace(r"[\$,]", "", regex=True).replace("", "0")
    )
    df["mrr"] = pd.to_numeric(df["mrr"], errors="coerce").fillna(0)
    df["payment_status"] = df["payment_status"].fillna("Unknown")
    print(f"  cleaned: {before} -> {len(df)} rows (removed {before - len(df)} duplicates)")
    return df


# ---------------------------------------------------------------- 2. KPIs
def compute_kpis(df: pd.DataFrame) -> dict:
    weeks = sorted(df["week"].unique())
    latest, prior = weeks[-1], weeks[-2]
    cur = df[df["week"] == latest]
    prev = df[df["week"] == prior]

    cur_rev, prev_rev = cur["mrr"].sum(), prev["mrr"].sum()
    wow = (cur_rev - prev_rev) / prev_rev * 100 if prev_rev else 0

    flagged = cur[cur["payment_status"].isin(["Late", "Overdue"])]
    top_ind = cur.groupby("industry")["mrr"].sum().idxmax()

    return {
        "latest_week": latest,
        "prior_week": prior,
        "total_mrr": round(cur_rev, 2),
        "wow_change_pct": round(wow, 1),
        "active_clients": int(cur["active_flag"].sum()),
        "churned_clients": int(cur["churned"].sum()),
        "new_clients": int(cur["new_client"].sum()),
        "late_or_overdue": int(len(flagged)),
        "revenue_at_risk": round(flagged["mrr"].sum(), 2),
        "top_industry": top_ind,
        "flagged_table": flagged[["client_id", "industry", "plan_type",
                                  "mrr", "payment_status"]].copy(),
    }


# ---------------------------------------------------------------- 3. LLM SUMMARY
def ai_summary(k: dict) -> str:
    """Ask Gemini to write an executive summary. Falls back to a template
    if no API key is set, so the pipeline always produces output."""
    facts = (
        f"Week: {k['latest_week']} (vs {k['prior_week']}). "
        f"Total MRR: ${k['total_mrr']:,.0f}, week-on-week {k['wow_change_pct']:+.1f}%. "
        f"Active clients: {k['active_clients']}. New: {k['new_clients']}. "
        f"Churned: {k['churned_clients']}. "
        f"Clients with late/overdue payments: {k['late_or_overdue']}, "
        f"totalling ${k['revenue_at_risk']:,.0f} at risk. "
        f"Top industry by revenue: {k['top_industry']}."
    )
    api_key = os.environ.get("GEMINI_API_KEY", "").strip()
    if api_key:
        prompt = (
            "You are an operations analyst. Write a concise 3-4 sentence executive "
            "summary of this week's client metrics for a leadership team. Be direct, "
            "highlight risks, and suggest one action. Data: " + facts
        )
        url = ("https://generativelanguage.googleapis.com/v1beta/models/"
               "gemini-2.0-flash:generateContent?key=" + api_key)
        body = json.dumps({"contents": [{"parts": [{"text": prompt}]}]}).encode()
        req = urllib.request.Request(url, data=body,
                                     headers={"Content-Type": "application/json"})
        try:
            with urllib.request.urlopen(req, timeout=30) as r:
                out = json.load(r)
            return out["candidates"][0]["content"]["parts"][0]["text"].strip()
        except Exception as e:
            print("  LLM call failed, using template. Reason:", e)

    # fallback template (still a real, useful summary)
    trend = "up" if k["wow_change_pct"] >= 0 else "down"
    return (
        f"In {k['latest_week']}, total recurring revenue was "
        f"${k['total_mrr']:,.0f}, {trend} {abs(k['wow_change_pct'])}% versus "
        f"{k['prior_week']}. The base held {k['active_clients']} active clients with "
        f"{k['new_clients']} newly onboarded and {k['churned_clients']} churned. "
        f"{k['late_or_overdue']} clients are late or overdue, placing "
        f"${k['revenue_at_risk']:,.0f} of monthly revenue at risk — collections should "
        f"prioritise these accounts. {k['top_industry']} remains the top revenue segment."
    )


# ---------------------------------------------------------------- 4. EXCEL REPORT
def build_report(k: dict, summary: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Weekly Report"

    navy = PatternFill("solid", fgColor="1F3864")
    grey = PatternFill("solid", fgColor="D9E1F2")
    white_bold = Font(name="Arial", bold=True, color="FFFFFF", size=14)
    hdr = Font(name="Arial", bold=True, color="FFFFFF")
    base = Font(name="Arial", size=11)
    thin = Border(*[Side(style="thin", color="BFBFBF")] * 4)

    ws.merge_cells("A1:E1")
    ws["A1"] = f"Weekly Operations Report — {k['latest_week']}"
    ws["A1"].fill = navy
    ws["A1"].font = white_bold
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # KPI block
    kpis = [
        ("Total MRR", f"${k['total_mrr']:,.0f}"),
        ("Week-on-Week", f"{k['wow_change_pct']:+.1f}%"),
        ("Active Clients", k["active_clients"]),
        ("New Clients", k["new_clients"]),
        ("Churned", k["churned_clients"]),
        ("Late / Overdue", k["late_or_overdue"]),
        ("Revenue at Risk", f"${k['revenue_at_risk']:,.0f}"),
        ("Top Industry", k["top_industry"]),
    ]
    r = 3
    for label, val in kpis:
        ws[f"A{r}"] = label
        ws[f"A{r}"].fill = grey
        ws[f"A{r}"].font = Font(name="Arial", bold=True)
        ws[f"B{r}"] = val
        ws[f"B{r}"].font = base
        r += 1

    # AI summary block
    ws[f"A{r+1}"] = "Executive Summary (AI-generated)"
    ws[f"A{r+1}"].font = Font(name="Arial", bold=True, size=12)
    ws.merge_cells(f"A{r+2}:E{r+5}")
    cell = ws[f"A{r+2}"]
    cell.value = summary
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    cell.font = base

    # flagged clients table
    t = r + 7
    ws[f"A{t}"] = "Clients Requiring Follow-Up (Late / Overdue)"
    ws[f"A{t}"].font = Font(name="Arial", bold=True, size=12)
    t += 1
    headers = ["Client ID", "Industry", "Plan", "MRR", "Payment Status"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=t, column=c, value=h)
        cell.fill = navy
        cell.font = hdr
        cell.border = thin
    for _, row in k["flagged_table"].iterrows():
        t += 1
        vals = [row["client_id"], row["industry"], row["plan_type"],
                round(row["mrr"], 2), row["payment_status"]]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=t, column=c, value=v)
            cell.font = base
            cell.border = thin

    for col, w in zip("ABCDE", [22, 16, 12, 14, 16]):
        ws.column_dimensions[col].width = w

    wb.save(REPORT_FILE)
    print(f"  saved {REPORT_FILE}")


# ---------------------------------------------------------------- MAIN
def main():
    print("1) Loading raw data...")
    df = pd.read_csv(RAW_FILE)
    print(f"   {len(df)} raw rows")

    print("2) Cleaning & validating...")
    df = clean(df)

    print("3) Computing KPIs...")
    k = compute_kpis(df)
    print(f"   MRR ${k['total_mrr']:,.0f} ({k['wow_change_pct']:+.1f}% WoW), "
          f"{k['late_or_overdue']} flagged, ${k['revenue_at_risk']:,.0f} at risk")

    print("4) Generating AI executive summary...")
    summary = ai_summary(k)
    with open(SUMMARY_FILE, "w") as f:
        f.write(summary)
    print("   ---\n  ", summary, "\n   ---")

    print("5) Building Excel report...")
    build_report(k, summary)
    print("\nDone. Open", REPORT_FILE)


if __name__ == "__main__":
    main()
